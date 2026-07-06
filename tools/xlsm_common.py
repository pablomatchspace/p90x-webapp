"""Shared parsing helpers for the P90Xcel workbook (v2.05).

Used by gen_catalog.py (one-time asset generation, outputs committed) and
convert_xlsm.py (personal data conversion, output gitignored). Both scripts
read fixed, position-based structures documented in PRD Appendix B and fail
loudly on anything unrecognized rather than guessing.
"""

from __future__ import annotations

import datetime as dt
import re
from dataclasses import dataclass, field

import openpyxl
from openpyxl.utils import get_column_letter

STRENGTH_SHEETS = {
    "CHEST & BACK": "chest-back",
    "SHOULDERS & ARMS": "shoulders-arms",
    "LEGS & BACK": "legs-back",
    "CORE SYNERGISTICS": "core-synergistics",
    "CORE SYNERGISTICS (LEAN)": "core-synergistics",  # same workout, Lean column capacity
    "CHEST, SHOULDERS, & TRICEPS": "chest-shoulders-triceps",
    "BACK & BICEPS": "back-biceps",
}

COMPLETION_SHEETS = {
    "PLYOMETRICS": "plyometrics",
    "CARDIO X": "cardio-x",
    "YOGA X": "yoga-x",
    "KENPO X": "kenpo-x",
}

ARX_SHEET = "AB RIPPER X"

# Schedule column-N text → workout keys. Keys must exist in the catalog.
SCHEDULE_NAME_MAP = {
    "Chest & Back, Ab Ripper": ["chest-back", "ab-ripper-x"],
    "Plyometrics": ["plyometrics"],
    "Shoulders & Arms, Ab Ripper": ["shoulders-arms", "ab-ripper-x"],
    "Yoga X": ["yoga-x"],
    "Legs & Back, Ab Ripper": ["legs-back", "ab-ripper-x"],
    "Kenpo X": ["kenpo-x"],
    "Rest or X Stretch": ["rest"],
    "X Strech": ["x-stretch"],  # workbook typo, corrected display name in catalog
    "X Stretch": ["x-stretch"],
    "Core Synergistics": ["core-synergistics"],
    "Chest, Shoulders, Triceps, Ab Ripper": ["chest-shoulders-triceps", "ab-ripper-x"],
    "Back & Biceps, Ab Ripper": ["back-biceps", "ab-ripper-x"],
    "Cardio X": ["cardio-x"],
}

# Display-name fixes for known workbook typos (converter maps by position, not name).
NAME_FIXES = {
    "Lawnmovers": "Lawnmowers",
}

SCHEDULE_FIRST_ROW = 16
SCHEDULE_MAX_ROW = 964
PROGRAM_DAYS = 90

EXERCISE_RE = re.compile(r"^(\d{2})\s+(.+)$")
# In N-column formulas the human-readable name is the last string literal of a
# HYPERLINK(...) call: HYPERLINK("#'SHEET'!C5","Name")
HYPERLINK_NAME_RE = re.compile(r'HYPERLINK\("[^"]*",\s*"([^"]+)"\)')
LEAN_IF_RE = re.compile(r'^=?IF\(SETUP!\$C\$7="Lean",(.*)\)$', re.DOTALL)


def load_books(path: str):
    """(values, formulas) workbook pair."""
    wb_v = openpyxl.load_workbook(path, data_only=True, keep_vba=False)
    wb_f = openpyxl.load_workbook(path, data_only=False, keep_vba=False)
    return wb_v, wb_f


def slugify(name: str) -> str:
    s = name.lower()
    s = s.replace("&", "and").replace("/", " ").replace("'", "")
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s


def cell(ws, row: int, col: int):
    return ws.cell(row=row, column=col).value


def as_date(value) -> str | None:
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    return None


def split_lean_if(formula: str) -> tuple[str, str] | None:
    """Return (lean_branch, classic_branch) source text of an IF(Lean, a, b)."""
    m = LEAN_IF_RE.match(formula.strip())
    if not m:
        return None
    inner = m.group(1)
    # split at the top-level comma separating the two branches
    depth = 0
    in_str = False
    for i, ch in enumerate(inner):
        if ch == '"':
            in_str = not in_str
        elif not in_str:
            if ch == "(":
                depth += 1
            elif ch == ")":
                depth -= 1
            elif ch == "," and depth == 0:
                return inner[:i], inner[i + 1 :]
    raise ValueError(f"Could not split Lean IF branches: {formula}")


def workout_name_from_branch(branch: str) -> str:
    branch = branch.strip()
    m = HYPERLINK_NAME_RE.search(branch)
    if m:
        return m.group(1)
    if branch.startswith('"') and branch.endswith('"'):
        return branch[1:-1]
    raise ValueError(f"Cannot extract workout name from: {branch}")


@dataclass
class ScheduleDay:
    row: int
    date: str
    classic: list[str] | None  # workout keys, None for a skipped (blank) day
    lean: list[str] | None


@dataclass
class ScheduleExtract:
    days: list[ScheduleDay] = field(default_factory=list)  # program rows incl. blanks
    skip_dates: list[str] = field(default_factory=list)
    body_rows: list[dict] = field(default_factory=list)


def _normalize_name(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", name.lower())


_NORMALIZED_NAME_MAP = {_normalize_name(k): v for k, v in SCHEDULE_NAME_MAP.items()}


def map_schedule_name(name: str) -> list[str]:
    """Punctuation/spacing-insensitive lookup — the workbook's N column is not
    perfectly consistent (e.g. 'Chest, Shoulders Triceps, Ab Ripper')."""
    key = _normalize_name(name)
    if key not in _NORMALIZED_NAME_MAP:
        raise ValueError(f"Unknown workout name in SCHEDULE: {name!r}")
    return _NORMALIZED_NAME_MAP[key]


def extract_schedule(wb_v, wb_f) -> ScheduleExtract:
    """Walk SCHEDULE rows: program days (col N), skips (blank N mid-program), body log."""
    ws_v = wb_v["SCHEDULE"]
    ws_f = wb_f["SCHEDULE"]
    out = ScheduleExtract()
    program_days_seen = 0
    row = SCHEDULE_FIRST_ROW
    while row <= SCHEDULE_MAX_ROW:
        date = as_date(cell(ws_v, row, 1))
        if date is None:
            break

        # --- body log columns: B weight, E BF%, G water, H bone, L zone-minutes
        weight = cell(ws_v, row, 2)
        body_fat = cell(ws_v, row, 5)
        water = cell(ws_v, row, 7)
        bone = cell(ws_v, row, 8)
        zone = cell(ws_v, row, 12)
        if any(isinstance(v, (int, float)) for v in (weight, body_fat, water, bone, zone)):
            out.body_rows.append(
                {
                    "date": date,
                    "weight": weight if isinstance(weight, (int, float)) else None,
                    "bodyFat": body_fat if isinstance(body_fat, (int, float)) else None,
                    "water": water if isinstance(water, (int, float)) else None,
                    "bone": bone if isinstance(bone, (int, float)) else None,
                    "zoneMinutes": zone if isinstance(zone, (int, float)) else None,
                }
            )

        # --- program column N
        if program_days_seen < PROGRAM_DAYS:
            n_value = cell(ws_v, row, 14)
            n_formula = cell(ws_f, row, 14)
            if n_value is None and n_formula is None:
                # blank day inside the program span → an applied "skip"
                out.days.append(ScheduleDay(row=row, date=date, classic=None, lean=None))
                out.skip_dates.append(date)
            else:
                if isinstance(n_formula, str) and n_formula.startswith("="):
                    branches = split_lean_if(n_formula)
                    if branches is not None:
                        lean_name = workout_name_from_branch(branches[0])
                        classic_name = workout_name_from_branch(branches[1])
                    else:
                        m = HYPERLINK_NAME_RE.search(n_formula)
                        if not m:
                            raise ValueError(f"Row {row}: unparseable N formula {n_formula!r}")
                        lean_name = classic_name = m.group(1)
                else:
                    lean_name = classic_name = str(n_value)
                out.days.append(
                    ScheduleDay(
                        row=row,
                        date=date,
                        classic=map_schedule_name(classic_name),
                        lean=map_schedule_name(lean_name),
                    )
                )
                program_days_seen += 1
        row += 1

    if program_days_seen != PROGRAM_DAYS:
        raise ValueError(f"Expected {PROGRAM_DAYS} program days, found {program_days_seen}")
    return out


@dataclass
class ExerciseDef:
    """One logged exercise line.

    rounds     number of entry rows (1, 2, or 4 — Strip-Set Curls)
    secondary  meaning of the second value per row:
               'weight' (R×W) | 'knee'/'chair' (assisted reps, count 1/chairFactor)
               | 'extra' (second count added in full: other side / bonus reps)
               | None (single count)
    agg        how the workbook aggregates rounds into the score: 'avg' | 'sum'
    labels     per-round display labels, e.g. [{main:'N1',secondary:'K1'}, …]
    """

    id: str
    name: str
    rounds: int
    secondary: str | None
    agg: str
    labels: list[dict]
    rows: list[int]  # sheet rows of each round (block 1)


def detect_blocks(ws_v) -> list[int]:
    """Start columns of the 5-wide week blocks (non-empty row-4 cells)."""
    cols = []
    for col in range(2, ws_v.max_column + 1):
        if cell(ws_v, 4, col) is not None:
            cols.append(col)
    if not cols:
        raise ValueError(f"{ws_v.title}: no week blocks found on row 4")
    return cols


def parse_exercises(ws_v, ws_f) -> list[ExerciseDef]:
    """Infer each exercise's entry structure from block 1 (labels in cols B/D, score in F)."""
    defs: list[ExerciseDef] = []
    exercise_rows: list[tuple[int, str]] = []
    for row in range(6, ws_v.max_row + 1):
        a = cell(ws_v, row, 1)
        if isinstance(a, str):
            m = EXERCISE_RE.match(a.strip())
            if m:
                exercise_rows.append((row, m.group(2).strip()))

    for idx, (row, raw_name) in enumerate(exercise_rows):
        end = exercise_rows[idx + 1][0] if idx + 1 < len(exercise_rows) else ws_v.max_row + 1
        rows = [
            r
            for r in range(row, end)
            if isinstance(cell(ws_v, r, 2), str) and cell(ws_v, r, 2).strip()
        ]
        if len(rows) not in (1, 2, 4):
            raise ValueError(f"{ws_v.title} {raw_name!r}: unexpected label rows {rows}")

        def label(r: int, col: int) -> str | None:
            v = cell(ws_v, r, col)
            return str(v).strip() if isinstance(v, str) and str(v).strip() else None

        b1 = label(rows[0], 2)
        d1 = label(rows[0], 4)
        name = NAME_FIXES.get(raw_name, raw_name)

        if d1 in ("W", "W1"):
            secondary = "weight"
        elif d1 in ("K", "K1"):
            secondary = "knee"
        elif d1 in ("C", "C1") and b1 in ("NC", "NC1"):
            secondary = "chair"
        elif d1 is not None:
            secondary = "extra"  # other side / bonus count, added in full
        else:
            secondary = None

        f_formula = cell(ws_f, rows[0], 6)
        agg = (
            "avg"
            if isinstance(f_formula, str) and "AVERAGE" in f_formula.upper()
            else "sum"
        )

        labels = []
        for r in rows:
            entry: dict = {"main": label(r, 2)}
            d = label(r, 4)
            if d is not None:
                entry["secondary"] = d
            labels.append(entry)

        defs.append(
            ExerciseDef(
                id=slugify(name),
                name=name,
                rounds=len(rows),
                secondary=secondary,
                agg=agg,
                labels=labels,
                rows=rows,
            )
        )
    if not defs:
        raise ValueError(f"{ws_v.title}: no exercises found")
    return defs


def detect_completion_blocks(ws_v) -> list[dict]:
    """Cardio-style sheets: anchor on 'COMPLETED?' rows.

    Returns [{week_row, date_row, status_row, notes_row, cols: [col,...]}]
    where cols are columns having a week label above the anchor.
    """
    blocks = []
    for row in range(1, ws_v.max_row + 1):
        a = cell(ws_v, row, 1)
        if isinstance(a, str) and a.strip().upper() == "COMPLETED?":
            week_row = row - 2
            cols = [
                col for col in range(2, ws_v.max_column + 1) if cell(ws_v, week_row, col) is not None
            ]
            if not cols:
                raise ValueError(f"{ws_v.title}: COMPLETED? row {row} has no week labels")
            blocks.append(
                {
                    "week_row": week_row,
                    "date_row": row - 1,
                    "status_row": row,
                    "notes_row": row + 1,
                    "cols": cols,
                }
            )
    if not blocks:
        raise ValueError(f"{ws_v.title}: no completion blocks found")
    return blocks


def arx_columns(ws_v) -> list[int]:
    """AB RIPPER X: session columns are those with a date on row 6."""
    return [col for col in range(2, ws_v.max_column + 1) if as_date(cell(ws_v, 6, col)) is not None]


def arx_exercises(ws_v) -> list[tuple[int, str]]:
    """(row, name) for the 11 ARX moves (rows 7..17, column A)."""
    out = []
    for row in range(7, 18):
        a = cell(ws_v, row, 1)
        if isinstance(a, str) and a.strip():
            out.append((row, a.strip()))
    if len(out) != 11:
        raise ValueError(f"AB RIPPER X: expected 11 exercises, found {len(out)}")
    return out


def column_letter(col: int) -> str:
    return get_column_letter(col)
